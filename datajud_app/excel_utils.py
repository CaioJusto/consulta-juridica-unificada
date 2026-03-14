"""Leitura de planilhas e montagem do arquivo de retorno."""

from __future__ import annotations

import json
import re
from collections.abc import Iterable
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

PROCESS_NUMBER_RE = re.compile(r"\D+")


def workbook_sheet_names(file_bytes: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True)
    return workbook.sheetnames


def read_sheet_rows(file_bytes: bytes, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook[sheet_name]

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [_normalize_header(cell, index) for index, cell in enumerate(rows[0], start=1)]
    data_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(value not in (None, "") for value in row):
            continue
        item = {
            headers[index]: row[index] if index < len(row) else None
            for index in range(len(headers))
        }
        data_rows.append(item)

    return headers, data_rows


def normalize_process_number(raw_value: Any) -> str:
    if raw_value is None:
        return ""
    if isinstance(raw_value, float) and raw_value.is_integer():
        raw_value = int(raw_value)
    return PROCESS_NUMBER_RE.sub("", str(raw_value).strip())


def flatten_process(source: dict[str, Any]) -> dict[str, Any]:
    assuntos = source.get("assuntos") or []
    movimentos = source.get("movimentos") or []
    primeiro_andamento = _first_movement(movimentos)
    andamento_mais_recente = _latest_movement(movimentos)
    andamento_resumos = [_movement_summary(movimento) for movimento in movimentos]
    assuntos_codigos = [str(item.get("codigo")) for item in assuntos if item.get("codigo")]
    assuntos_nomes = [item.get("nome", "") for item in assuntos if item.get("nome")]
    tipos_andamento = _distinct_ordered(
        movement.get("nome", "") for movement in movimentos if movement.get("nome")
    )

    return {
        "id": source.get("id"),
        "numero_processo": source.get("numeroProcesso"),
        "tribunal": source.get("tribunal"),
        "timestamp_indice": source.get("@timestamp"),
        "grau": source.get("grau"),
        "nivel_sigilo": source.get("nivelSigilo"),
        "data_ajuizamento": source.get("dataAjuizamento"),
        "ultima_atualizacao": source.get("dataHoraUltimaAtualizacao"),
        "classe": _nested(source, "classe", "nome"),
        "classe_codigo": _nested(source, "classe", "codigo"),
        "orgao_julgador": _nested(source, "orgaoJulgador", "nome"),
        "orgao_julgador_codigo": _nested(source, "orgaoJulgador", "codigo"),
        "codigo_municipio_ibge": _nested(source, "orgaoJulgador", "codigoMunicipioIBGE"),
        "sistema": _nested(source, "sistema", "nome"),
        "sistema_codigo": _nested(source, "sistema", "codigo"),
        "formato": _nested(source, "formato", "nome"),
        "formato_codigo": _nested(source, "formato", "codigo"),
        "quantidade_assuntos": len(assuntos),
        "assuntos": ", ".join(assuntos_nomes),
        "assuntos_codigos": ", ".join(assuntos_codigos),
        "quantidade_andamentos": len(movimentos),
        "primeiro_andamento": primeiro_andamento.get("nome") if primeiro_andamento else None,
        "data_primeiro_andamento": (
            primeiro_andamento.get("dataHora") if primeiro_andamento else None
        ),
        "ultimo_andamento": andamento_mais_recente.get("nome") if andamento_mais_recente else None,
        "data_ultimo_andamento": (
            andamento_mais_recente.get("dataHora") if andamento_mais_recente else None
        ),
        "tipos_de_andamento": ", ".join(tipos_andamento),
        "andamentos": " | ".join(item for item in andamento_resumos if item),
        "json_bruto": json.dumps(source, ensure_ascii=False),
    }


def movements_rows(processes: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source in processes:
        for index, movement in enumerate(source.get("movimentos") or [], start=1):
            complementos = movement.get("complementosTabelados") or []
            rows.append(
                {
                    "numero_processo": source.get("numeroProcesso"),
                    "tribunal": source.get("tribunal"),
                    "ordem": index,
                    "codigo": movement.get("codigo"),
                    "andamento": movement.get("nome"),
                    "data_hora": movement.get("dataHora"),
                    "quantidade_complementos": len(complementos),
                    "orgao_julgador_codigo": _nested(movement, "orgaoJulgador", "codigo"),
                    "orgao_julgador": _nested(movement, "orgaoJulgador", "nome"),
                    "complementos": ", ".join(
                        _complemento_label(complemento)
                        for complemento in complementos
                        if _complemento_label(complemento)
                    ),
                    "json_bruto": json.dumps(movement, ensure_ascii=False),
                }
            )
    return rows


def movement_complements_rows(processes: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source in processes:
        for movement_index, movement in enumerate(source.get("movimentos") or [], start=1):
            for complement_index, complement in enumerate(
                movement.get("complementosTabelados") or [], start=1
            ):
                rows.append(
                    {
                        "numero_processo": source.get("numeroProcesso"),
                        "tribunal": source.get("tribunal"),
                        "ordem_andamento": movement_index,
                        "andamento": movement.get("nome"),
                        "data_hora_andamento": movement.get("dataHora"),
                        "ordem_complemento": complement_index,
                        "tipo": complement.get("descricao"),
                        "nome": complement.get("nome"),
                        "codigo": complement.get("codigo"),
                        "valor": complement.get("valor"),
                        "json_bruto": json.dumps(complement, ensure_ascii=False),
                    }
                )
    return rows


def subjects_rows(processes: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source in processes:
        for index, subject in enumerate(source.get("assuntos") or [], start=1):
            rows.append(
                {
                    "numero_processo": source.get("numeroProcesso"),
                    "tribunal": source.get("tribunal"),
                    "ordem": index,
                    "codigo": subject.get("codigo"),
                    "assunto": subject.get("nome"),
                    "json_bruto": json.dumps(subject, ensure_ascii=False),
                }
            )
    return rows


def raw_rows(processes: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "numero_processo": source.get("numeroProcesso"),
            "tribunal": source.get("tribunal"),
            "json_bruto": json.dumps(source, ensure_ascii=False),
        }
        for source in processes
    ]


def build_result_workbook(
    *,
    merged_rows: list[dict[str, Any]],
    process_rows: list[dict[str, Any]],
    movement_rows: list[dict[str, Any]],
    movement_complement_rows: list[dict[str, Any]],
    subject_rows: list[dict[str, Any]],
    credit_analysis_rows: list[dict[str, Any]] | None = None,
    processual_process_rows: list[dict[str, Any]] | None = None,
    processual_party_rows: list[dict[str, Any]] | None = None,
    processual_event_rows: list[dict[str, Any]] | None = None,
    processual_distribution_rows: list[dict[str, Any]] | None = None,
    processual_petition_rows: list[dict[str, Any]] | None = None,
    processual_document_rows: list[dict[str, Any]] | None = None,
    processual_incident_rows: list[dict[str, Any]] | None = None,
    official_consulta_rows: list[dict[str, Any]] | None = None,
    official_process_rows: list[dict[str, Any]] | None = None,
    official_party_rows: list[dict[str, Any]] | None = None,
    official_lawyer_rows: list[dict[str, Any]] | None = None,
    official_event_rows: list[dict[str, Any]] | None = None,
    official_document_rows: list[dict[str, Any]] | None = None,
    raw_process_rows: list[dict[str, Any]],
    not_found_rows: list[dict[str, Any]],
) -> bytes:
    official_consulta_rows = official_consulta_rows or []
    official_process_rows = official_process_rows or []
    official_party_rows = official_party_rows or []
    official_lawyer_rows = official_lawyer_rows or []
    official_event_rows = official_event_rows or []
    official_document_rows = official_document_rows or []
    credit_analysis_rows = credit_analysis_rows or []
    processual_process_rows = processual_process_rows or []
    processual_party_rows = processual_party_rows or []
    processual_event_rows = processual_event_rows or []
    processual_distribution_rows = processual_distribution_rows or []
    processual_petition_rows = processual_petition_rows or []
    processual_document_rows = processual_document_rows or []
    processual_incident_rows = processual_incident_rows or []

    workbook = Workbook()
    first_sheet = workbook.active
    workbook.remove(first_sheet)

    _write_sheet(workbook, "consulta_mesclada", merged_rows)
    _write_sheet(workbook, "processos", process_rows)
    _write_sheet(workbook, "movimentos", movement_rows)
    _write_sheet(workbook, "complementos_mov", movement_complement_rows)
    _write_sheet(workbook, "assuntos", subject_rows)
    _write_sheet(workbook, "analise_credito", credit_analysis_rows)
    _write_sheet(workbook, "trf1_processual", processual_process_rows)
    _write_sheet(workbook, "trf1_proc_partes", processual_party_rows)
    _write_sheet(workbook, "trf1_proc_movs", processual_event_rows)
    _write_sheet(workbook, "trf1_proc_dist", processual_distribution_rows)
    _write_sheet(workbook, "trf1_proc_pet", processual_petition_rows)
    _write_sheet(workbook, "trf1_proc_docs", processual_document_rows)
    _write_sheet(workbook, "trf1_proc_incs", processual_incident_rows)
    _write_sheet(workbook, "consulta_processo", official_consulta_rows)
    _write_sheet(workbook, "fontes_oficiais", official_process_rows)
    _write_sheet(workbook, "partes_oficiais", official_party_rows)
    _write_sheet(workbook, "advogados_oficiais", official_lawyer_rows)
    _write_sheet(workbook, "eventos_oficiais", official_event_rows)
    _write_sheet(workbook, "documentos_oficiais", official_document_rows)
    _write_sheet(workbook, "json_completo", raw_process_rows)
    _write_sheet(workbook, "nao_encontrados", not_found_rows)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_named_sheets_workbook(sheet_rows: list[tuple[str, list[dict[str, Any]]]]) -> bytes:
    workbook = Workbook()
    first_sheet = workbook.active
    workbook.remove(first_sheet)

    for title, rows in sheet_rows:
        _write_sheet(workbook, title, rows)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def default_download_name(prefix: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.xlsx"


def _write_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title=title)

    if not rows:
        worksheet.append(["sem_dados"])
        return

    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 60)


def _nested(data: dict[str, Any], *keys: str) -> Any:
    current: Any = data
    for key in keys:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def _normalize_header(value: Any, index: int) -> str:
    text = str(value).strip() if value not in (None, "") else f"coluna_{index}"
    return text


def _latest_movement(movements: list[dict[str, Any]]) -> dict[str, Any] | None:
    with_date = [movement for movement in movements if movement.get("dataHora")]
    if not with_date:
        return movements[-1] if movements else None
    return max(with_date, key=lambda movement: movement.get("dataHora", ""))


def _first_movement(movements: list[dict[str, Any]]) -> dict[str, Any] | None:
    with_date = [movement for movement in movements if movement.get("dataHora")]
    if not with_date:
        return movements[0] if movements else None
    return min(with_date, key=lambda movement: movement.get("dataHora", ""))


def _movement_summary(movement: dict[str, Any]) -> str:
    parts = [
        str(movement.get("dataHora") or "").strip(),
        str(movement.get("nome") or "").strip(),
        str(_nested(movement, "orgaoJulgador", "nome") or "").strip(),
    ]
    complementos = movement.get("complementosTabelados") or []
    complemento_text = ", ".join(
        _complemento_label(complemento)
        for complemento in complementos
        if _complemento_label(complemento)
    )
    if complemento_text:
        parts.append(complemento_text)
    return " - ".join(part for part in parts if part)


def _complemento_label(complemento: dict[str, Any]) -> str:
    nome = str(complemento.get("nome") or complemento.get("descricao") or "").strip()
    valor = str(complemento.get("valor") or "").strip()
    if nome and valor:
        return f"{nome}: {valor}"
    return nome or valor


def _distinct_ordered(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out
